import sys
import re
from collections import defaultdict, deque
from datetime import datetime
from json import loads
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

OUTPUT_XLSX = Path("hp_release_registry.xlsx")
COLUMNS = ["No", "Release Name", "Group", "Label", "Release Date", "Release Type", "Edition", "Disc Number",
           "Disc Order", "Track Number", "Track Name", "Singer", "Lyricist", "Composer", "Arranger"]
KEY_COLUMNS = ["Disc Number", "Disc Order", "Track Number"]


def _str(value):
    return "" if value is None else str(value).strip()


def save_excel(rows):
    workbook = load_workbook(OUTPUT_XLSX) if OUTPUT_XLSX.exists() else Workbook()
    sheet = workbook.worksheets[0]
    sheet.title = "Releases"

    headers = [_str(sheet.cell(row=1, column=i).value) for i in range(1, sheet.max_column + 1)]
    if not any(headers):
        headers = COLUMNS[:]
    headers += [header for header in COLUMNS if header not in headers]
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=i, value=header)

    col = {header: i for i, header in enumerate(headers, start=1)}
    existing_rows = defaultdict(deque)
    for i in range(2, sheet.max_row + 1):
        key = tuple(_str(sheet.cell(row=i, column=col[header]).value) for header in KEY_COLUMNS)
        if any(key):
            existing_rows[key].append(i)

    appended = 0
    for row in rows:
        key = tuple(_str(row[header]) for header in KEY_COLUMNS)
        if existing_rows[key]:
            excel_row = existing_rows[key].popleft()
            for header in COLUMNS[1:]:
                if (excel_value := _str(sheet.cell(row=excel_row, column=col[header]).value)) != (
                fetched_value := _str(row[header])):
                    print(
                        f"[Warning] Existing row differs "
                        f"(Disc Number={key[0]}, Disc Order={key[1]}, Track Number={key[2]}, Column={header}): "
                        f"Excel={excel_value!r}, Fetched={fetched_value!r}",
                        file=sys.stderr,
                    )
            continue

        row["No"] = sheet.max_row
        appended += 1
        sheet.append([row.get(header) for header in headers])

    workbook.save(OUTPUT_XLSX)
    return appended


def main():
    release_blocks = []
    seen_release_links = set()
    for domain in ["https://helloproject.com", "https://up-front-works.jp"]:
        release_page = BeautifulSoup(requests.get(f"{domain}/release/").text, "lxml")
        version_dir = loads(str(release_page.find("astro-island").get("props")))["versionDir"][1]
        years = list(map(lambda v: v[1], loads(str(release_page.find("astro-island").get("props")))["years"][1]))
        print(f"Latest release version: {version_dir}")
        print(f"Release years: {years}")
        for year in years:
            # if year != "2020":
            # if int(year) < 2024:
            #     continue
            print(f"\n=== {year} Releases ===")
            year_json = requests.get(f"{domain}/json/{version_dir}/{year}_releases.json").json()
            for release in year_json["items"]:
                if "single" not in release["category"] and "album" not in release["category"]:
                    continue
                print([release["category"], release["title"], release["link"].split("/")[2]])
                if release["link"] in seen_release_links:
                    continue
                seen_release_links.add(release["link"])
                detail_page = BeautifulSoup(requests.get(f"{domain}" + release["link"]).text, "lxml")
                release_title = detail_page.select_one("h1.ReleaseHead__mainName").text.strip()
                release_group = detail_page.select_one("div.ReleaseHead__mainTitle > div:nth-of-type(2)").text.strip()
                release_date_string = detail_page.select_one(
                    "div.ReleaseHead__mainDetails > dl:nth-of-type(1) dd").text.strip()
                release_date = datetime.strptime(release_date_string, "%Y.%m.%d")
                label = detail_page.select_one("div.ReleaseHead__mainDetails > dl:nth-of-type(2) dd").text.strip()
                release_type = "single" if "single" in release["category"] else "album"
                release_rows = []
                for release_edition in detail_page.select("div.ReleaseEdition"):
                    edition = edition[0].text.strip() if (
                        edition := release_edition.select("div.ReleaseEdition__name")) else ""
                    disk_id = re.search(
                        r"[A-Z]+-\d+",
                        release_edition.select_one("div.TrackList div.ReleaseEdition__headline").get_text(" ", strip=True),
                    ).group()
                    id_struct = re.match(r"([A-Z]+)-(\d+)", disk_id)
                    disk_order = 0
                    for track_list_index, disk in enumerate(release_edition.select("div.TrackList")):
                        if disk.select_one("div.ReleaseEdition__mediaType").text.strip() != "CD":
                            continue
                        disk_order += 1
                        disk_number = (
                            f"{id_struct.group(1)}-"
                            f"{int(id_struct.group(2)) + track_list_index:0{len(id_struct.group(2))}d}"
                        )
                        for track_number, track in enumerate(disk.select("div.TrackListItem"), start=1):
                            track_title = track.select_one("div.TrackListItem__title").text.strip()
                            lyricist = None
                            composer = None
                            arranger = None
                            singer = None
                            for note in track.select("div.TrackListItem__notes span"):
                                if note.text.strip().startswith("作詞："):
                                    lyricist = note.text.strip()[3:]
                                elif note.text.strip().startswith("作曲："):
                                    composer = note.text.strip()[3:]
                                elif note.text.strip().startswith("編曲："):
                                    arranger = note.text.strip()[3:]
                                elif note.text.strip().startswith("歌："):
                                    singer = note.text.strip()[2:]
                            release_rows.append({
                                "No": None,
                                "Release Name": release_title,
                                "Group": release_group,
                                "Label": label,
                                "Release Date": release_date.strftime("%Y/%m/%d"),
                                "Release Type": release_type,
                                "Edition": edition,
                                "Disc Number": disk_number,
                                "Disc Order": disk_order,
                                "Track Number": track_number,
                                "Track Name": track_title,
                                "Singer": singer,
                                "Lyricist": lyricist,
                                "Composer": composer,
                                "Arranger": arranger,
                            })
                if release_rows:
                    release_blocks.append((release_date, release_rows))

    release_blocks.sort(key=lambda block: block[0])
    rows = [row for _, release_rows in release_blocks for row in release_rows]
    appended = save_excel(rows)
    print(f"Fetched {len(rows)} rows. Appended {appended} rows to {OUTPUT_XLSX}.")


if __name__ == '__main__':
    main()
